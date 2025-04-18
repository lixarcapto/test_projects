

import openpyxl

def read_nested_column_xlsx(
        ROUTE:str)->dict:
    """
    función de persistencia que lee un 
    archivo Excel como un diccionario 
    anidado en horizontal
    # no se porque funciona no entiendo
    # openpyxl
    """
    workbook = openpyxl.load_workbook(ROUTE)
    worksheet = workbook.active
    key_list = []
    for row in worksheet.iter_rows(min_row=2, min_col=1, max_col=1):
        key_list.append(row[0].value)
    # carga las demas columnas
    diccionario_columnas = {}
    for col_num in range(2, worksheet.max_column + 1):
        columna = worksheet[col_num - 1]
        valores_columna = []
        for row in worksheet.iter_rows(min_row=2, min_col=col_num, max_col=col_num):
            valores_columna.append(row[0].value)
        diccionario_columnas[worksheet.cell(row=1, column=col_num).value] = valores_columna
    final_dict = {}
    for k in diccionario_columnas:
       final_dict[k] = dict(zip(
               key_list, 
               diccionario_columnas[k]
        ))
    return final_dict