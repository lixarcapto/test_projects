
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from typing import Dict, Any

def write_xlsx_flat_dict(
        DICT: Dict[Any, Any], 
        PATH: str):
    """
    Escribe los pares clave-valor de un diccionario en un archivo .xlsx.
    Cada par se escribe en una nueva fila, con la clave en la columna A y el valor en la B.

    Args:
        diccionario (Dict[Any, Any]): El diccionario con los datos a escribir.
        ruta_archivo (str): La ruta completa donde se guardará el archivo .xlsx.
    """
    if not isinstance(DICT, dict):
        print("❌ Error: La entrada no es un diccionario válido.")
        return

    try:
        # 1. Crear un nuevo libro de trabajo (workbook)
        workbook: Workbook = openpyxl.Workbook()
        
        # 2. Seleccionar la hoja de trabajo activa
        hoja_activa: Worksheet = workbook.active
        hoja_activa.title = "Datos del Diccionario"

        # 3. Inicializar el contador de filas
        fila = 1
        
        # 4. Iterar a través de los pares clave-valor del diccionario
        for clave, valor in DICT.items():
            # Escribir la clave en la primera columna (columna 'A')
            hoja_activa.cell(row=fila, column=1, value=clave)
            
            # Escribir el valor en la segunda columna (columna 'B')
            hoja_activa.cell(row=fila, column=2, value=valor)
            
            # Avanzar a la siguiente fila
            fila += 1
            
        # 5. Guardar el archivo Excel
        workbook.save(PATH)
        print(f"✅ Diccionario guardado exitosamente en '{PATH}'")

    except Exception as e:
        print(f"❌ Ocurrió un error al intentar escribir el archivo Excel: {e}")