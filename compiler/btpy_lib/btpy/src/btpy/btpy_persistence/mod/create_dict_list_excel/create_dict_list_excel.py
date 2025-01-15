



from openpyxl import Workbook

def create_dict_list_excel(
        DICT_LIST:dict[list], PATH:str)\
        ->None:
    """
    funcion que reciba un diccionario 
    de listas y almacene en un archivo 
    excel como elementos de la primera 
    columna las claves del diccionario, 
    y los elementos de la lista de cada 
    clave deben deben escribirse en la 
    fila de clave.
    """
    path:str = PATH
    if(not ".xlsx" in PATH):
        path += ".xlsx"
    workbook = Workbook()
    hoja = workbook.active
    fila = 1
    for clave, valor in DICT_LIST.items():
        hoja.cell(row=fila, column=1, value=clave)
        for i, elemento in enumerate(valor, start=2):
            hoja.cell(row=fila, column=i, value=elemento)
        fila += 1
    workbook.save(path)