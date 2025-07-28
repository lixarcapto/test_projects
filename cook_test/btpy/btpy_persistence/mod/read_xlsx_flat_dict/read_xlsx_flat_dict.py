


from openpyxl import load_workbook

def read_xlsx_flat_dict(PATH_XLSX:str)\
        ->dict:
    """
     Función que lee todos los pares 
     clave valor que se encuentren en un 
     archivo Excel como un diccionario.   
     leerá cada primera celda de la fila 
     video de la segunda celda como un 
     par clave valor y así sucesivamente 
     con todas las celdas.
    """
    sheet_name=None
    # Cargar el archivo Excel
    wb = load_workbook(PATH_XLSX)
    # Seleccionar la hoja de trabajo (si no se especifica, usa la hoja activa)
    ws = wb[sheet_name] if sheet_name else wb.active
    # Inicializar un diccionario para almacenar los resultados
    result_dict = {}
    # Recorrer las filas y convertir 
    # cada una en un diccionario
    leng = 0
    for row in ws.iter_rows(values_only=True):
        leng = len(row)
        for x in range(0, leng,
             2):
            key = row[x]  # La primera celda es la clave
            if(key == None): continue
            values = row[x +1]  # Las celdas siguientes son los valores
            result_dict[key] = values
    return result_dict