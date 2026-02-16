
from openpyxl import Workbook

def write_xlsx_nested_dict_row(
        PATH:str, 
        NESTED_DICT:dict):
    
    # 1. Crear el libro y la hoja activa
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte"

    # 2. Obtener todos los encabezados únicos (columnas) de los diccionarios internos
    # Usamos una lista para mantener el orden de aparición
    columnas = []
    for sub_dict in NESTED_DICT.values():
        for clave in sub_dict.keys():
            if clave not in columnas:
                columnas.append(clave)

    # 3. Escribir los encabezados en la fila 1
    # La columna 1 será para la "Clave de Fila" (el ID del diccionario superior)
    ws.cell(row=1, column=1, value="ID / Clave")
    
    for col_idx, nombre_columna in enumerate(columnas, start=2):
        ws.cell(row=1, column=col_idx, value=nombre_columna)

    # 4. Escribir los datos fila por fila
    for fila_idx, (id_fila, contenido) in enumerate(NESTED_DICT.items(), start=2):
        # Escribimos el ID en la primera columna
        ws.cell(row=fila_idx, column=1, value=id_fila)
        
        # Escribimos cada valor en su columna correspondiente
        for col_idx, nombre_columna in enumerate(columnas, start=2):
            # Usamos .get() por si a alguna fila le falta esa clave específica
            valor = contenido.get(nombre_columna, "")
            ws.cell(row=fila_idx, column=col_idx, value=valor)

    # 5. Guardar el archivo
    wb.save(PATH)
    print(f"Archivo '{PATH}' creado con éxito.")