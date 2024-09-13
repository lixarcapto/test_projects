


import openpyxl


def create_dict_list_xlsx(data:dict[list], 
        filename:str)->None:
    """
    Stores a dictionary of arrays in an 
    Excel file.
    """
    FORMAT = ".xlsx"
    # Create a new Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write the dictionary keys to the 
    # first column of the Excel file
    for i, key in enumerate(data.keys()):
        ws.cell(row=i + 1, column=1)\
            .value = key
    # Write the array elements to the 
    # corresponding cells in the Excel file
    for i, (key, values) in enumerate(
            data.items()):
        for j, value in enumerate(values):
            ws.cell(row=i + 1, column=j + 2)\
                .value = value
    # Save the Excel file
    wb.save(filename + FORMAT)